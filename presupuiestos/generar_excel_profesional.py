#!/usr/bin/env python3
"""
Generador de Presupuesto Excel Profesional
Chatbot CocoLuVentas - Con fórmulas automáticas
Autor: Ember Drago
"""

import sys
import os

# Intentar importar openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    print("✅ Usando openpyxl")
    TIENE_OPENPYXL = True
except ImportError:
    print("⚠️  openpyxl no disponible")
    print("📦 Instalando openpyxl...")
    
    # Intentar instalar
    import subprocess
    try:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', '--user', 'openpyxl', '-q'])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter
        print("✅ openpyxl instalado exitosamente")
        TIENE_OPENPYXL = True
    except Exception as e:
        print(f"❌ No se pudo instalar openpyxl: {e}")
        TIENE_OPENPYXL = False

if not TIENE_OPENPYXL:
    print("\n🔧 SOLUCIÓN:")
    print("1. Instala openpyxl manualmente:")
    print("   sudo emerge dev-python/openpyxl")
    print("   O descarga desde: https://pypi.org/project/openpyxl/")
    print("\n2. O usa este CSV y ábrelo con LibreOffice:")
    print("   libreoffice --calc Presupuesto_CocoLuVentas.csv")
    sys.exit(1)

# ============================================================================
# CONFIGURACIÓN DEL PRESUPUESTO
# ============================================================================

# Tarifa por hora (USD)
TARIFA_HORA = 50

# Tareas del proyecto [nombre, descripcion, horas]
TAREAS = {
    'FASE 1: ANÁLISIS Y DISEÑO': [
        ('1.1 Análisis requerimientos', 'Reunión con cliente, definición flujos', 4),
        ('1.2 Diseño arquitectura', 'Stack tecnológico, estructura', 6),
        ('1.3 Diseño base datos', 'Modelado vendedores, historial', 4),
        ('1.4 Flujos conversacionales', 'Mapeo comandos y respuestas', 6),
    ],
    'FASE 2: CONFIGURACIÓN ENTORNO': [
        ('2.1 Setup Node.js', 'Instalación BuilderBot, Baileys', 8),
        ('2.2 Configuración repositorio', 'Git, estructura carpetas', 2),
        ('2.3 Configuración WhatsApp', 'Baileys provider, autenticación', 4),
    ],
    'FASE 3: DESARROLLO CORE': [
        ('3.1 Sistema vendedores', '5 vendedores con datos', 4),
        ('3.2 Rotación Round-Robin', 'Algoritmo asignación secuencial', 6),
        ('3.3 Flow asignación', 'Keywords y asignación automática', 8),
        ('3.4 Flow lista vendedores', 'Comando ver equipo', 3),
        ('3.5 Flow estadísticas', 'Tracking asignaciones', 5),
        ('3.6 Historial', 'Registro temporal', 4),
    ],
    'FASE 4: FUNCIONALIDADES AVANZADAS': [
        ('4.1 Respuestas personalizadas', 'Templates con datos', 4),
        ('4.2 Múltiples comandos', 'Keywords variadas', 3),
        ('4.3 Sistema logs', 'Registro consola', 3),
        ('4.4 Manejo errores', 'Try-catch, fallbacks', 4),
    ],
    'FASE 5: INTEGRACIÓN API REST': [
        ('5.1 API REST básica', 'Endpoints stats y mensajes', 8),
        ('5.2 Endpoint estadísticas', 'GET /v1/stats', 3),
        ('5.3 Endpoint mensajes', 'POST /v1/messages', 4),
        ('5.4 Documentación API', 'README con ejemplos', 4),
    ],
    'FASE 6: TESTING Y QA': [
        ('6.1 Pruebas unitarias', 'Tests funciones', 12),
        ('6.2 Pruebas integración', 'Tests flujos', 10),
        ('6.3 Pruebas WhatsApp real', 'QR, envío/recepción', 8),
        ('6.4 Testing rotación', 'Verificar Round-Robin', 4),
        ('6.5 Pruebas estrés', 'Mensajes simultáneos', 6),
        ('6.6 Corrección bugs', 'Fix problemas', 12),
    ],
    'FASE 7: DOCUMENTACIÓN': [
        ('7.1 Manual usuario', 'Guía comandos', 5),
        ('7.2 Guía instalación', 'Scripts, instrucciones', 5),
        ('7.3 Documentación técnica', 'Arquitectura, código', 6),
        ('7.4 README completo', 'GitHub con ejemplos', 4),
    ],
    'FASE 8: DEPLOYMENT': [
        ('8.1 Scripts inicio', 'Auto-inicio', 4),
        ('8.2 Config producción', 'Variables, seguridad', 6),
        ('8.3 Training cliente', 'Capacitación', 4),
        ('8.4 Entrega final', 'Transferencia', 4),
    ],
    'FASE 9: SOPORTE (1 MES)': [
        ('9.1 Soporte técnico', 'Resolución problemas', 20),
        ('9.2 Ajustes menores', 'Modificaciones', 10),
        ('9.3 Monitoreo', 'Verificación', 6),
    ],
}

# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def crear_estilos():
    """Crear estilos reutilizables"""
    # Colores
    azul_oscuro = "366092"
    azul_claro = "DCE6F1"
    naranja = "FFC000"
    verde = "92D050"
    
    # Bordes
    borde_delgado = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    return {
        'header': {
            'fill': PatternFill(start_color=azul_oscuro, end_color=azul_oscuro, fill_type='solid'),
            'font': Font(color='FFFFFF', bold=True, size=12),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': borde_delgado
        },
        'section': {
            'fill': PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid'),
            'font': Font(bold=True, size=11),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': borde_delgado
        },
        'total': {
            'fill': PatternFill(start_color=naranja, end_color=naranja, fill_type='solid'),
            'font': Font(bold=True, size=14),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': borde_delgado
        },
        'normal': {
            'alignment': Alignment(vertical='center', wrap_text=True),
            'border': borde_delgado
        }
    }

def aplicar_estilo(cell, estilo_dict):
    """Aplicar estilo a una celda"""
    for attr, value in estilo_dict.items():
        setattr(cell, attr, value)

# ============================================================================
# GENERAR EXCEL
# ============================================================================

def generar_excel():
    print("\n🚀 Generando Excel profesional...\n")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    
    # Obtener estilos
    estilos = crear_estilos()
    
    # ========================================================================
    # ENCABEZADO
    # ========================================================================
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = 'PRESUPUESTO - CHATBOT WHATSAPP COCOLUVENTAS'
    cell.font = Font(size=18, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # Info del proyecto
    ws['A3'] = 'Cliente:'
    ws['B3'] = 'CocoLuVentas'
    ws['A4'] = 'Proyecto:'
    ws['B4'] = 'Sistema de Chatbot WhatsApp con Rotación de Vendedores'
    ws['A5'] = 'Fecha:'
    ws['B5'] = '=TODAY()'
    ws['B5'].number_format = 'DD/MM/YYYY'
    ws['A6'] = 'Desarrollador:'
    ws['B6'] = 'Ember Drago'
    
    # TARIFA AJUSTABLE
    ws['A7'] = 'TARIFA/HORA:'
    ws['B7'] = TARIFA_HORA
    ws['B7'].number_format = '"$"#,##0'
    ws['B7'].font = Font(bold=True, size=14, color='FF0000')
    ws['C7'] = '← AJUSTA AQUÍ PARA RECALCULAR TODO'
    ws['C7'].font = Font(italic=True, color='FF0000')
    
    for row in [3, 4, 5, 6, 7]:
        ws[f'A{row}'].font = Font(bold=True)
    
    # ========================================================================
    # TABLA DE TAREAS
    # ========================================================================
    
    # Headers
    row = 9
    headers = ['Fase/Tarea', 'Descripción', 'Horas', 'Días', 'Costo/Hora', 'Total USD', '% Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        aplicar_estilo(cell, estilos['header'])
    
    # Tareas
    row_inicial_tareas = row + 1
    current_row = row + 1
    filas_para_sumar = []
    
    for fase, tareas in TAREAS.items():
        # Fila de fase
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = fase
        aplicar_estilo(cell, estilos['section'])
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Tareas de la fase
        for tarea, descripcion, horas in tareas:
            filas_para_sumar.append(current_row)
            
            # Columna A: Tarea
            cell = ws.cell(row=current_row, column=1, value=tarea)
            aplicar_estilo(cell, estilos['normal'])
            
            # Columna B: Descripción
            cell = ws.cell(row=current_row, column=2, value=descripcion)
            aplicar_estilo(cell, estilos['normal'])
            
            # Columna C: Horas
            cell = ws.cell(row=current_row, column=3, value=horas)
            aplicar_estilo(cell, estilos['normal'])
            cell.number_format = '#,##0'
            
            # Columna D: Días (fórmula: horas/8)
            cell = ws.cell(row=current_row, column=4, value=f'=C{current_row}/8')
            aplicar_estilo(cell, estilos['normal'])
            cell.number_format = '0.00'
            
            # Columna E: Costo/Hora (referencia a B7)
            cell = ws.cell(row=current_row, column=5, value='=$B$7')
            aplicar_estilo(cell, estilos['normal'])
            cell.number_format = '"$"#,##0'
            
            # Columna F: Total (fórmula: horas * costo/hora)
            cell = ws.cell(row=current_row, column=6, value=f'=C{current_row}*E{current_row}')
            aplicar_estilo(cell, estilos['normal'])
            cell.number_format = '"$"#,##0'
            
            # Columna G: % (se llenará después)
            cell = ws.cell(row=current_row, column=7, value='')
            aplicar_estilo(cell, estilos['normal'])
            
            current_row += 1
        
        # Espacio entre fases
        current_row += 1
    
    # ========================================================================
    # FILA DE TOTALES
    # ========================================================================
    row_total = current_row
    ws.merge_cells(f'A{row_total}:B{row_total}')
    cell = ws[f'A{row_total}']
    cell.value = 'TOTAL PROYECTO'
    aplicar_estilo(cell, estilos['total'])
    
    # Total horas
    rangos_horas = '+'.join([f'C{r}' for r in filas_para_sumar])
    cell = ws[f'C{row_total}']
    cell.value = f'={rangos_horas}'
    aplicar_estilo(cell, estilos['total'])
    cell.number_format = '#,##0'
    
    # Total días
    cell = ws[f'D{row_total}']
    cell.value = f'=C{row_total}/8'
    aplicar_estilo(cell, estilos['total'])
    cell.number_format = '0.0'
    
    # Costo/hora
    ws.merge_cells(f'E{row_total}:E{row_total}')
    
    # Total USD
    rangos_total = '+'.join([f'F{r}' for r in filas_para_sumar])
    cell = ws[f'F{row_total}']
    cell.value = f'={rangos_total}'
    aplicar_estilo(cell, estilos['total'])
    cell.number_format = '"$"#,##0'
    cell.font = Font(bold=True, size=16, color='FF0000')
    
    # Llenar porcentajes
    for r in filas_para_sumar:
        cell = ws[f'G{r}']
        cell.value = f'=F{r}/$F${row_total}*100'
        cell.number_format = '0.0"%"'
    
    # ========================================================================
    # RESUMEN EJECUTIVO
    # ========================================================================
    row = row_total + 3
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = 'RESUMEN EJECUTIVO'
    cell.font = Font(size=16, bold=True)
    cell.fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 30
    
    row += 2
    resumen = [
        ('Total de Horas:', f'=C{row_total}', 'horas', '#,##0.0'),
        ('Total de Días (8h/día):', f'=D{row_total}', 'días laborables', '#,##0.0'),
        ('Semanas (5 días/sem):', f'=D{row_total}/5', 'semanas', '#,##0.0'),
        ('Duración con Imprevistos (+30%):', f'=D{row_total}*1.3', 'días calendario', '#,##0.0'),
        ('Tarifa por Hora:', '=$B$7', 'USD', '"$"#,##0'),
        ('COSTO TOTAL PROYECTO:', f'=F{row_total}', 'USD', '"$"#,##0'),
    ]
    
    for titulo, formula, unidad, formato in resumen:
        ws[f'A{row}'] = titulo
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = formato
        if 'COSTO TOTAL' in titulo:
            ws[f'B{row}'].font = Font(bold=True, size=16, color='FF0000')
            ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'C{row}'] = unidad
        ws[f'C{row}'].font = Font(italic=True)
        row += 1
    
    # ========================================================================
    # HOJA 2: COMPARACIÓN DE ESCENARIOS
    # ========================================================================
    ws2 = wb.create_sheet("Comparación")
    
    ws2.merge_cells('A1:F1')
    ws2['A1'] = 'COMPARACIÓN DE ESCENARIOS'
    ws2['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws2['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 35
    
    # Headers
    row = 3
    for col, header in enumerate(['Escenario', 'Horas', 'Días', 'Semanas', 'Costo USD', '% vs Base'], 1):
        cell = ws2.cell(row=row, column=col, value=header)
        aplicar_estilo(cell, estilos['header'])
    
    # Escenarios
    escenarios = [
        ('Optimista (-20%)', 0.8, 'DCE6F1'),
        ('REALISTA (Base)', 1.0, 'FFC000'),
        ('Conservador (+30%)', 1.3, 'F4B084'),
        ('Pesimista (+50%)', 1.5, 'FF6B6B'),
    ]
    
    for idx, (nombre, factor, color) in enumerate(escenarios, 4):
        ws2[f'A{idx}'] = nombre
        ws2[f'A{idx}'].font = Font(bold=True)
        ws2[f'A{idx}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        ws2[f'B{idx}'] = f'=Presupuesto!$C${row_total}*{factor}'
        ws2[f'B{idx}'].number_format = '#,##0'
        
        ws2[f'C{idx}'] = f'=B{idx}/8'
        ws2[f'C{idx}'].number_format = '#,##0.0'
        
        ws2[f'D{idx}'] = f'=C{idx}/5'
        ws2[f'D{idx}'].number_format = '#,##0.0'
        
        ws2[f'E{idx}'] = f'=Presupuesto!$B$7*B{idx}'
        ws2[f'E{idx}'].number_format = '"$"#,##0'
        
        ws2[f'F{idx}'] = f'={factor-1}*100'
        ws2[f'F{idx}'].number_format = '+0.0"% ";-0.0"% ";0"%"'
    
    # Anchos de columna
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 12
    
    # ========================================================================
    # HOJA 3: DESGLOSE POR FASE
    # ========================================================================
    ws3 = wb.create_sheet("Desglose por Fase")
    
    ws3.merge_cells('A1:E1')
    ws3['A1'] = 'DESGLOSE POR FASE'
    ws3['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws3['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws3['A1'].alignment = Alignment(horizontal='center')
    ws3.row_dimensions[1].height = 35
    
    # Headers
    for col, header in enumerate(['Fase', 'Horas', 'Costo USD', '% del Total', 'Días'], 1):
        cell = ws3.cell(row=3, column=col, value=header)
        aplicar_estilo(cell, estilos['header'])
    
    # Calcular rangos por fase
    fase_row = 4
    fase_actual_row = row_inicial_tareas
    
    for fase_nombre in TAREAS.keys():
        # Encontrar filas de esta fase
        inicio_fase = fase_actual_row + 1
        num_tareas = len(TAREAS[fase_nombre])
        fin_fase = inicio_fase + num_tareas - 1
        
        ws3[f'A{fase_row}'] = fase_nombre
        ws3[f'A{fase_row}'].font = Font(bold=True)
        
        ws3[f'B{fase_row}'] = f'=SUM(Presupuesto!C{inicio_fase}:C{fin_fase})'
        ws3[f'B{fase_row}'].number_format = '#,##0'
        
        ws3[f'C{fase_row}'] = f'=SUM(Presupuesto!F{inicio_fase}:F{fin_fase})'
        ws3[f'C{fase_row}'].number_format = '"$"#,##0'
        
        ws3[f'D{fase_row}'] = f'=C{fase_row}/Presupuesto!$F${row_total}*100'
        ws3[f'D{fase_row}'].number_format = '0.0"%"'
        
        ws3[f'E{fase_row}'] = f'=B{fase_row}/8'
        ws3[f'E{fase_row}'].number_format = '#,##0.0'
        
        fase_row += 1
        fase_actual_row = fin_fase + 2  # +2 por fila vacía entre fases
    
    # Anchos
    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 12
    
    # ========================================================================
    # AJUSTAR ANCHOS EN HOJA PRINCIPAL
    # ========================================================================
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    
    # ========================================================================
    # GUARDAR ARCHIVO
    # ========================================================================
    output_file = 'Presupuesto_CocoLuVentas_PROFESIONAL.xlsx'
    wb.save(output_file)
    
    print(f"✅ Excel creado exitosamente!")
    print(f"📊 Archivo: {output_file}")
    print(f"📁 Ubicación: {os.path.abspath(output_file)}")
    print(f"\n💡 INSTRUCCIONES:")
    print(f"   1. Abre el archivo con LibreOffice Calc o Excel")
    print(f"   2. Ajusta la TARIFA/HORA en celda B7")
    print(f"   3. ¡TODO se recalcula automáticamente!")
    print(f"\n📋 Incluye 3 hojas:")
    print(f"   • Presupuesto - Desglose completo con fórmulas")
    print(f"   • Comparación - Escenarios optimista/realista/conservador")
    print(f"   • Desglose por Fase - Resumen por cada fase")
    print(f"\n🎯 Total actual: {sum([t[2] for fase in TAREAS.values() for t in fase])} horas")
    print(f"   Costo: ${sum([t[2] for fase in TAREAS.values() for t in fase]) * TARIFA_HORA:,}")
    
    return output_file

# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    try:
        archivo = generar_excel()
        print(f"\n🎉 ¡Listo! Abre: {archivo}")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
