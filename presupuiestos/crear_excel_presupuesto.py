#!/usr/bin/env python3
"""
Script para crear presupuesto Excel del Chatbot CocoLuVentas
con tiempos realistas de desarrollo
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Presupuesto Chatbot"

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
section_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
total_font = Font(bold=True, size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezado del documento
ws.merge_cells('A1:F1')
ws['A1'] = 'PRESUPUESTO - CHATBOT WHATSAPP COCOLUVENTAS'
ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Información del proyecto
ws['A3'] = 'Cliente:'
ws['B3'] = 'CocoLuVentas'
ws['A4'] = 'Proyecto:'
ws['B4'] = 'Sistema de Chatbot WhatsApp con Rotación de Vendedores'
ws['A5'] = 'Fecha:'
ws['B5'] = datetime.now().strftime('%d/%m/%Y')
ws['A6'] = 'Desarrollador:'
ws['B6'] = 'Ember Drago'

for row in [3, 4, 5, 6]:
    ws[f'A{row}'].font = Font(bold=True)

# Encabezados de tabla
row = 8
headers = ['Fase/Tarea', 'Descripción', 'Horas', 'Días', 'Costo/Hora', 'Total']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Tareas del proyecto con tiempos REALISTAS
tareas = [
    # FASE 1: Análisis y Diseño
    ('FASE 1: ANÁLISIS Y DISEÑO', '', '', '', '', ''),
    ('1.1 Análisis de requerimientos', 'Reunión con cliente, definición de flujos', 4, 0.5, 50, '=C10*E10'),
    ('1.2 Diseño de arquitectura', 'Definir stack tecnológico, estructura', 6, 0.75, 50, '=C11*E11'),
    ('1.3 Diseño de base de datos', 'Modelado de datos, vendedores, historial', 4, 0.5, 50, '=C12*E12'),
    ('1.4 Diseño de flujos conversacionales', 'Mapeo de comandos y respuestas', 6, 0.75, 50, '=C13*E13'),
    
    # FASE 2: Configuración del Entorno
    ('FASE 2: CONFIGURACIÓN DEL ENTORNO', '', '', '', '', ''),
    ('2.1 Setup Node.js y dependencias', 'Instalación de BuilderBot, Baileys', 3, 0.4, 50, '=C15*E15'),
    ('2.2 Configuración de repositorio', 'Git, estructura de carpetas', 2, 0.25, 50, '=C16*E16'),
    ('2.3 Configuración de WhatsApp Web', 'Baileys provider, autenticación', 4, 0.5, 50, '=C17*E17'),
    
    # FASE 3: Desarrollo del Core
    ('FASE 3: DESARROLLO DEL CORE', '', '', '', '', ''),
    ('3.1 Sistema de vendedores', 'Configuración de 5 vendedores con datos', 4, 0.5, 50, '=C19*E19'),
    ('3.2 Sistema de rotación Round-Robin', 'Algoritmo de asignación secuencial', 6, 0.75, 50, '=C20*E20'),
    ('3.3 Flow de asignación automática', 'Detección de keywords, asignación', 8, 1, 50, '=C21*E21'),
    ('3.4 Flow de lista de vendedores', 'Comando para ver equipo completo', 3, 0.4, 50, '=C22*E22'),
    ('3.5 Flow de estadísticas', 'Tracking de asignaciones, reportes', 5, 0.6, 50, '=C23*E23'),
    ('3.6 Historial de asignaciones', 'Registro temporal de asignaciones', 4, 0.5, 50, '=C24*E24'),
    
    # FASE 4: Funcionalidades Avanzadas
    ('FASE 4: FUNCIONALIDADES AVANZADAS', '', '', '', '', ''),
    ('4.1 Respuestas personalizadas', 'Templates con datos del vendedor', 4, 0.5, 50, '=C26*E26'),
    ('4.2 Manejo de múltiples comandos', 'Keywords variadas, sinónimos', 3, 0.4, 50, '=C27*E27'),
    ('4.3 Sistema de logs', 'Registro en consola de actividad', 3, 0.4, 50, '=C28*E28'),
    ('4.4 Validación y manejo de errores', 'Try-catch, fallbacks', 4, 0.5, 50, '=C29*E29'),
    
    # FASE 5: Integración y API
    ('FASE 5: INTEGRACIÓN Y API REST', '', '', '', '', ''),
    ('5.1 API REST básica', 'Endpoints para stats y mensajes', 6, 0.75, 50, '=C31*E31'),
    ('5.2 Endpoint de estadísticas', 'GET /v1/stats', 3, 0.4, 50, '=C32*E32'),
    ('5.3 Endpoint de envío de mensajes', 'POST /v1/messages', 4, 0.5, 50, '=C33*E33'),
    ('5.4 Documentación de API', 'README con ejemplos de uso', 3, 0.4, 50, '=C34*E34'),
    
    # FASE 6: Testing y QA
    ('FASE 6: TESTING Y ASEGURAMIENTO DE CALIDAD', '', '', '', '', ''),
    ('6.1 Pruebas unitarias', 'Test de funciones principales', 8, 1, 50, '=C36*E36'),
    ('6.2 Pruebas de integración', 'Test de flujos completos', 6, 0.75, 50, '=C37*E37'),
    ('6.3 Pruebas con WhatsApp real', 'QR scanning, envío/recepción', 6, 0.75, 50, '=C38*E38'),
    ('6.4 Testing de rotación', 'Verificar Round-Robin funciona', 4, 0.5, 50, '=C39*E39'),
    ('6.5 Pruebas de estrés', 'Múltiples mensajes simultáneos', 4, 0.5, 50, '=C40*E40'),
    ('6.6 Corrección de bugs', 'Fix de problemas encontrados', 8, 1, 50, '=C41*E41'),
    
    # FASE 7: Documentación
    ('FASE 7: DOCUMENTACIÓN', '', '', '', '', ''),
    ('7.1 Manual de usuario', 'Guía de comandos para clientes', 4, 0.5, 50, '=C43*E43'),
    ('7.2 Guía de instalación', 'Scripts e instrucciones de setup', 5, 0.6, 50, '=C44*E44'),
    ('7.3 Documentación técnica', 'Arquitectura, código, APIs', 6, 0.75, 50, '=C45*E45'),
    ('7.4 README completo', 'GitHub README con ejemplos', 3, 0.4, 50, '=C46*E46'),
    
    # FASE 8: Deployment
    ('FASE 8: DEPLOYMENT Y ENTREGA', '', '', '', '', ''),
    ('8.1 Preparación de scripts', 'Scripts de inicio automático', 4, 0.5, 50, '=C48*E48'),
    ('8.2 Configuración de entorno producción', 'Variables de entorno, seguridad', 5, 0.6, 50, '=C49*E49'),
    ('8.3 Entrenamiento al cliente', 'Sesión de capacitación', 4, 0.5, 50, '=C50*E50'),
    ('8.4 Entrega y documentación final', 'Transferencia de archivos', 3, 0.4, 50, '=C51*E51'),
    
    # FASE 9: Soporte Post-Lanzamiento
    ('FASE 9: SOPORTE POST-LANZAMIENTO (1 MES)', '', '', '', '', ''),
    ('9.1 Soporte técnico', 'Resolución de dudas y problemas', 20, 2.5, 50, '=C53*E53'),
    ('9.2 Ajustes menores', 'Modificaciones de vendedores, textos', 8, 1, 50, '=C54*E54'),
    ('9.3 Monitoreo de estabilidad', 'Verificación de funcionamiento', 6, 0.75, 50, '=C55*E55'),
]

# Agregar tareas
current_row = row + 1
for tarea in tareas:
    for col, valor in enumerate(tarea, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = valor
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Estilo para secciones
        if tarea[1] == '' and 'FASE' in str(tarea[0]):
            cell.fill = section_fill
            cell.font = section_font
            if col == 1:
                ws.merge_cells(f'{get_column_letter(col)}{current_row}:{get_column_letter(6)}{current_row}')
    
    current_row += 1

# Totales
total_row = current_row + 1
ws.merge_cells(f'A{total_row}:B{total_row}')
ws[f'A{total_row}'] = 'TOTAL PROYECTO'
ws[f'A{total_row}'].font = total_font
ws[f'A{total_row}'].fill = total_fill
ws[f'A{total_row}'].alignment = Alignment(horizontal='center', vertical='center')

ws[f'C{total_row}'] = f'=SUMIF(C10:C{current_row-1},"<>",C10:C{current_row-1})'
ws[f'C{total_row}'].font = total_font
ws[f'C{total_row}'].fill = total_fill
ws[f'C{total_row}'].number_format = '0'

ws[f'D{total_row}'] = f'=C{total_row}/8'
ws[f'D{total_row}'].font = total_font
ws[f'D{total_row}'].fill = total_fill
ws[f'D{total_row}'].number_format = '0.0'

ws.merge_cells(f'E{total_row}:E{total_row}')
ws[f'F{total_row}'] = f'=SUM(F10:F{current_row-1})'
ws[f'F{total_row}'].font = total_font
ws[f'F{total_row}'].fill = total_fill
ws[f'F{total_row}'].number_format = '$#,##0.00'

# Resumen ejecutivo
summary_row = total_row + 3
ws.merge_cells(f'A{summary_row}:F{summary_row}')
ws[f'A{summary_row}'] = 'RESUMEN EJECUTIVO'
ws[f'A{summary_row}'].font = Font(size=14, bold=True)
ws[f'A{summary_row}'].fill = section_fill
ws[f'A{summary_row}'].alignment = Alignment(horizontal='center')

summary_row += 1
resumen = [
    ('Total de Horas:', f'=C{total_row}', 'horas'),
    ('Total de Días Laborables:', f'=D{total_row}', 'días (8 hrs/día)'),
    ('Duración Estimada (calendario):', f'=D{total_row}*1.5', 'días (incluyendo imprevistos)'),
    ('Costo por Hora:', 50, 'USD'),
    ('COSTO TOTAL DEL PROYECTO:', f'=F{total_row}', 'USD'),
]

for item in resumen:
    ws[f'A{summary_row}'] = item[0]
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'] = item[1]
    if isinstance(item[1], str) and item[1].startswith('='):
        ws[f'B{summary_row}'].number_format = '#,##0.0'
    ws[f'C{summary_row}'] = item[2]
    summary_row += 1

# Notas
notes_row = summary_row + 2
ws.merge_cells(f'A{notes_row}:F{notes_row}')
ws[f'A{notes_row}'] = 'NOTAS IMPORTANTES'
ws[f'A{notes_row}'].font = Font(size=12, bold=True)
ws[f'A{notes_row}'].fill = section_fill

notes_row += 1
notas = [
    '• Los tiempos son estimados y pueden variar según complejidad de requerimientos',
    '• Incluye soporte post-lanzamiento de 1 mes',
    '• No incluye hosting o infraestructura de servidor',
    '• Requiere cuenta de WhatsApp Business del cliente',
    '• El proyecto está desarrollado con BuilderBot (Leifer Méndez)',
    '• Tecnología: Node.js, Baileys, WhatsApp Web Protocol',
    '• Sistema de rotación Round-Robin automatizado',
    '• 5 vendedores configurables con información personalizada',
]

for nota in notas:
    ws[f'A{notes_row}'] = nota
    ws.merge_cells(f'A{notes_row}:F{notes_row}')
    ws[f'A{notes_row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[notes_row].height = 20
    notes_row += 1

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 15

# Guardar archivo
output_file = '/home/guest/Documents/chatboot-cocoluventas/presupuiestos/Presupuesto_CocoLuVentas_Detallado.xlsx'
wb.save(output_file)
print(f'✅ Excel creado exitosamente: {output_file}')
print(f'📊 Revisa el archivo para análisis de tiempos')
